# app/domains/action/services/report_builder.py
import json
from pathlib import Path
from sqlalchemy.orm import Session
import openpyxl

from app.domains.action import repository
from app.domains.action.models import ReportFormat
from app.domains.action.mongo_repository import get_meeting_summary
from app.domains.action.services import thumbnail as thumb
from app.domains.action.services.wbs_builder import build_wbs_template

STORAGE_ROOT = Path("storage")

def _thumb_path(meeting_id: int, suffix: str) -> Path:
    p = STORAGE_ROOT / "meetings" / str(meeting_id)
    p.mkdir(parents=True, exist_ok=True)
    return p / f"thumb_{suffix}.webp"

def _report_file_path(meeting_id: int, report_id: int, ext: str) -> Path:
    p = STORAGE_ROOT / "meetings" / str(meeting_id) / "reports"
    p.mkdir(parents=True, exist_ok=True)
    return p / f"{report_id}.{ext}"

def generate_markdown(db: Session, meeting_id: int, user_id: int):
    minute = repository.get_meeting_minute(db, meeting_id)
    if not minute or not minute.content:
        raise ValueError("회의록이 없습니다.")
    
    meeting = repository.get_meeting(db, meeting_id)

    thumb_path = _thumb_path(meeting_id, "md")
    thumb.generate_text_thumbnail(minute.content, str(thumb_path))

    return repository.save_report(
        db, meeting_id, user_id,
        format=ReportFormat.markdown,
        title=f"{meeting.title} 회의록",
        content=minute.content,
        thumbnail_url=str(thumb_path),
    )

def generate_html(db: Session, meeting_id: int, user_id: int):
    minute = repository.get_meeting_minute(db, meeting_id)
    if not minute or not minute.content:
        raise ValueError("회의록이 없습니다.")
    
    meeting = repository.get_meeting(db, meeting_id)

    # 마크다운과 썸네일 같음
    thumb_path = _thumb_path(meeting_id, "md")
    if not thumb_path.exists():
        thumb.generate_text_thumbnail(minute.content, str(thumb_path))

    return repository.save_report(
        db=db,
        meeting_id=meeting_id,
        created_by=user_id,
        format=ReportFormat.html,
        title=f"{meeting.title} HTML 보고서",
        thumbnail_url=str(thumb_path),
    )

def generate_excel(db: Session, meeting_id: int, user_id: int):
    summary = get_meeting_summary(meeting_id)
    if not summary:
        raise ValueError("회의 요약 데이터가 없습니다.")
    
    meeting = repository.get_meeting(db, meeting_id)

    # ID 확보를 위해 먼저 row 생성
    report = repository.save_report(
        db=db,
        meeting_id=meeting_id,
        created_by=user_id,
        format=ReportFormat.excel,
        title=f"{meeting.title} Excel 보고서",
    )

    file_path = _report_file_path(meeting_id, report.id, "xlsx")
    thumb_path = _thumb_path(meeting_id, "excel")

    # Excel 생성
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "개요"
    overview = summary.get("overview", {})
    ws.append(["목적",   overview.get("purpose", "")])
    ws.append(["일시",   overview.get("datetime_str", "")])
    ws.append(["참석자", ", ".join(summary.get("attendees", []))])

    ws2 = wb.create_sheet("결정사항")
    ws2.append(["결정사항", "근거", "반대의견"])
    for d in summary.get("decisions", []):
        ws2.append([d.get("decision", ""), d.get("rationale", ""), d.get("opposing_opinion", "")])

    ws3 = wb.create_sheet("액션아이템")
    ws3.append(["담당자", "내용", "마감", "우선순위", "긴급도"])
    for a in summary.get("action_items", []):
        ws3.append([
            a.get("assignee", ""), 
            a.get("content", ""),
            a.get("deadline", ""), 
            a.get("priority", ""), 
            a.get("urgency", "")
        ])

    ws4 = wb.create_sheet("미결사항")
    ws4.append(["내용", "이월"])
    for p in summary.get("pending_items", []):
        ws4.append([p.get("content", ""), "O" if p.get("carried_over") else "X"])

    wb.save(str(file_path))
    thumb.generate_format_thumbnail("excel", str(thumb_path))

    report.file_url      = str(file_path)
    report.thumbnail_url = str(thumb_path)
    db.commit()
    db.refresh(report)
    return report


async def generate_wbs(db: Session, meeting_id: int, user_id: int):
    wbs     = await build_wbs_template(db, meeting_id)
    meeting = repository.get_meeting(db, meeting_id)

    thumb_path = _thumb_path(meeting_id, "wbs")
    thumb.generate_format_thumbnail("wbs", str(thumb_path))

    return repository.save_report(
        db, meeting_id, user_id,
        format=ReportFormat.wbs,
        title=f"{meeting.title} WBS",
        content=json.dumps(wbs, ensure_ascii=False),
        thumbnail_url=str(thumb_path),
    )